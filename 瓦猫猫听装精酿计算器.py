"""
瓦猫猫听装精酿计算器 - 门店点餐订单报表分析工具
统计各门店 瓦猫猫听装精酿 的出品数量
统计范围: 瓦猫猫听装(规格=听, 含二销套餐内听装) + 全部1L装(名称含1L)
最终结果: (听装出品数量 + 1L装出品数量) ÷ 12
"""

import re
import os
import io
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime

APP_FONT_FAMILY = "PingFang SC" if os.name == "posix" and hasattr(os, "uname") and os.uname().sysname == "Darwin" else "Microsoft YaHei"

# 排除的支付方式（与提成计算器基础排除一致）
EXCLUDED_PAYMENT_METHODS = [
    '打折支付', '礼品券兑换', '会员支付（赠送）', '赠送商品', '会员积分兑换'
]


def extract_store_name(filename: str) -> str:
    """从文件名提取门店名称"""
    match = re.search(r'【(.+?)】', filename)
    if match:
        return match.group(1)
    return os.path.splitext(filename)[0]


def load_csv(filepath: str) -> pd.DataFrame:
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            content = f.read().replace('\t', '')
        df = pd.read_csv(io.StringIO(content))
        df.columns = [col.strip() for col in df.columns]
        return df
    except Exception as e:
        raise Exception(f"加载文件失败: {e}")


def count_wamaomao_ting(df: pd.DataFrame):
    """统计瓦猫猫听装精酿的出品数量
    返回: (听装数量, 1L装数量, 合计数量)
    """
    base_filter = ~df['商品名称'].astype(str).str.contains(r'\[退\]', regex=True, na=False)
    base_filter = base_filter & ~df['支付方式'].astype(str).isin(EXCLUDED_PAYMENT_METHODS)
    # 商品分类剔除「招牌原浆精酿*」(原浆大桶/套餐不属于瓦猫猫听装精酿)
    base_filter = base_filter & ~df['商品分类'].astype(str).str.contains('招牌原浆精酿', na=False)
    base_df = df[base_filter].copy()

    # 听装: 商品名称包含瓦猫猫 且 商品规格为"听"（含二销套餐内的听装明细）
    ting_filter = base_df['商品名称'].astype(str).str.contains('瓦猫猫', na=False)
    ting_filter = ting_filter & (base_df['商品规格'].astype(str).str.strip() == '听')
    ting_count = int(base_df[ting_filter]['出品数量'].sum())

    # 1L装: 商品名称包含 1L 或（1L）的所有产品（含瓦猫猫1L、代号9等）
    l1_filter = base_df['商品名称'].astype(str).str.contains(r'[（(]1L[）)]', regex=True, na=False)
    l1_count = int(base_df[l1_filter]['出品数量'].sum())

    return ting_count, l1_count, ting_count + l1_count


class WamaomaoCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("瓦猫猫听装精酿计算器 - 门店点餐订单报表分析工具")
        self.root.geometry("1000x680")

        self.selected_files = []
        self.results = {}

        self.setup_ui()

    def setup_ui(self):
        title_label = tk.Label(
            self.root,
            text="瓦猫猫听装精酿计算器",
            font=(APP_FONT_FAMILY, 24, "bold"),
            fg="#2C3E50"
        )
        title_label.pack(pady=20)

        file_frame = ttk.LabelFrame(self.root, text="选择文件", padding=10)
        file_frame.pack(fill="x", padx=20, pady=10)

        self.select_btn = ttk.Button(
            file_frame,
            text="选择CSV文件",
            command=self.select_files
        )
        self.select_btn.pack(side="left", padx=5)

        self.clear_btn = ttk.Button(
            file_frame,
            text="清空",
            command=self.clear_files
        )
        self.clear_btn.pack(side="left", padx=5)

        self.calc_btn = ttk.Button(
            file_frame,
            text="计算",
            command=self.calculate,
            state="disabled"
        )
        self.calc_btn.pack(side="left", padx=5)

        self.export_btn = ttk.Button(
            file_frame,
            text="导出Excel",
            command=self.export_excel,
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=5)

        self.file_listbox = tk.Listbox(
            self.root,
            height=6,
            font=(APP_FONT_FAMILY, 10)
        )
        self.file_listbox.pack(fill="x", padx=20, pady=5)

        result_frame = ttk.LabelFrame(self.root, text="计算结果", padding=10)
        result_frame.pack(fill="both", expand=True, padx=20, pady=10)

        columns = ("门店", "听装出品数", "1L装出品数", "小计")
        self.result_tree = ttk.Treeview(
            result_frame,
            columns=columns,
            show="headings",
            height=12
        )

        self.result_tree.heading("门店", text="门店")
        self.result_tree.column("门店", width=280, anchor="center")
        self.result_tree.heading("听装出品数", text="听装出品数")
        self.result_tree.column("听装出品数", width=150, anchor="center")
        self.result_tree.heading("1L装出品数", text="1L装出品数")
        self.result_tree.column("1L装出品数", width=150, anchor="center")
        self.result_tree.heading("小计", text="小计")
        self.result_tree.column("小计", width=150, anchor="center")

        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar.set)

        self.result_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.status_var = tk.StringVar(value="请选择CSV文件")
        self.status_label = tk.Label(
            self.root,
            textvariable=self.status_var,
            font=(APP_FONT_FAMILY, 9),
            fg="#7F8C8D"
        )
        self.status_label.pack(pady=5)

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="选择点餐订单报表CSV文件",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )

        if files:
            self.selected_files = list(files)
            self.file_listbox.delete(0, tk.END)

            for f in self.selected_files:
                store_name = extract_store_name(os.path.basename(f))
                self.file_listbox.insert(tk.END, f"{store_name} - {os.path.basename(f)}")

            self.status_var.set(f"已选择 {len(files)} 个文件")
            self.calc_btn.config(state="normal")

    def clear_files(self):
        self.selected_files = []
        self.file_listbox.delete(0, tk.END)
        self.result_tree.delete(*self.result_tree.get_children())
        self.results = {}
        self.status_var.set("请选择CSV文件")
        self.calc_btn.config(state="disabled")
        self.export_btn.config(state="disabled")

    def calculate(self):
        if not self.selected_files:
            messagebox.showwarning("警告", "请先选择文件")
            return

        self.status_var.set("正在计算...")
        self.result_tree.delete(*self.result_tree.get_children())
        self.results = {}

        try:
            total_ting = 0
            total_l1 = 0
            for filepath in self.selected_files:
                store_name = extract_store_name(os.path.basename(filepath))
                self.status_var.set(f"正在处理: {store_name}")
                self.root.update()

                df = load_csv(filepath)
                ting_count, l1_count, sub_total = count_wamaomao_ting(df)
                self.results[store_name] = (ting_count, l1_count, sub_total)
                total_ting += ting_count
                total_l1 += l1_count

                self.result_tree.insert("", tk.END, values=(store_name, ting_count, l1_count, sub_total))

            grand_total = total_ting + total_l1
            result = grand_total / 12

            self.result_tree.insert("", tk.END, values=("合计", total_ting, total_l1, grand_total), tags=("total",))
            self.result_tree.tag_configure("total", background="#D3D3D3", font=(APP_FONT_FAMILY, 10, "bold"))

            self.status_var.set(
                f"计算完成，共处理 {len(self.results)} 个门店，"
                f"听装+1L装合计: {grand_total}，最终结果(÷12): {result:g}"
            )
            self.export_btn.config(state="normal")
            messagebox.showinfo("完成", f"计算完成！\n听装+1L装合计: {grand_total}\n最终结果(÷12): {result:g}")

        except Exception as e:
            self.status_var.set(f"计算失败: {str(e)}")
            messagebox.showerror("错误", f"计算失败: {str(e)}")

    def export_excel(self):
        if not self.results:
            messagebox.showwarning("警告", "没有可导出的数据")
            return

        filepath = filedialog.asksaveasfilename(
            title="保存Excel报告",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"瓦猫猫听装精酿报表_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "瓦猫猫听装精酿统计"

            ws['A1'] = "瓦猫猫听装精酿统计报表"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')

            headers = ["门店", "听装出品数", "1L装出品数", "小计"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(3, col_idx, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            total_ting = 0
            total_l1 = 0
            for row_idx, (store_name, (ting_count, l1_count, sub_total)) in enumerate(self.results.items(), 4):
                ws.cell(row_idx, 1, store_name)
                ws.cell(row_idx, 2, ting_count)
                ws.cell(row_idx, 3, l1_count)
                ws.cell(row_idx, 4, sub_total)
                total_ting += ting_count
                total_l1 += l1_count

            grand_total = total_ting + total_l1
            total_row = len(self.results) + 4
            ws.cell(total_row, 1, "合计").font = Font(bold=True)
            ws.cell(total_row, 2, total_ting)
            ws.cell(total_row, 3, total_l1)
            ws.cell(total_row, 4, grand_total).font = Font(bold=True)

            result_row = total_row + 2
            ws.cell(result_row, 1, "最终结果(÷12)")
            ws.cell(result_row, 1).font = Font(bold=True, size=12)
            ws.cell(result_row, 4, grand_total / 12).font = Font(bold=True, size=12, color="CC0000")

            for col, width in zip("ABCD", [30, 16, 16, 16]):
                ws.column_dimensions[col].width = width

            wb.save(filepath)
            self.status_var.set(f"已导出: {filepath}")
            messagebox.showinfo("完成", f"Excel报表已保存到:\n{filepath}")

        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")


def main():
    root = tk.Tk()
    app = WamaomaoCalculatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
