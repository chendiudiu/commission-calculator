"""
提成计算器 - 门店点餐订单报表分析工具
用于计算各门店的提成统计数据
"""

import re
import os
import io
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime


# ============ 核心计算逻辑 ============

# 排除的支付方式
EXCLUDED_PAYMENT_METHODS = [
    '打折支付', '礼品券兑换', '会员支付（赠送）', '赠送商品', '会员积分兑换'
]


def extract_store_name(filename: str) -> str:
    """从文件名提取门店名称"""
    match = re.search(r'【(.+?)】', filename)
    if match:
        return match.group(1)
    return os.path.splitext(filename)[0]


def load_csv(filepath: str) -> pd.DataFrame:
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            content = f.read().replace('\t', '')
        df = pd.read_csv(io.StringIO(content))
        df.columns = [col.strip() for col in df.columns]
        return df
    except Exception as e:
        raise Exception(f"加载文件失败: {e}")


def is_refund(row) -> bool:
    """检查是否为退单"""
    return '[退]' in str(row.get('商品名称', ''))


def is_excluded_payment(row) -> bool:
    """检查是否使用了排除的支付方式"""
    payment = str(row.get('支付方式', ''))
    return payment in EXCLUDED_PAYMENT_METHODS


def calculate_commission(df: pd.DataFrame) -> dict:
    """计算所有提成数据"""
    # 基础过滤：排除退单和排除的支付方式
    base_filter = ~df['商品名称'].astype(str).str.contains(r'\[退\]', regex=True, na=False)
    base_filter = base_filter & ~df['支付方式'].astype(str).isin(EXCLUDED_PAYMENT_METHODS)
    base_df = df[base_filter].copy()
    
    result = {}
    
    # ===== 方法1: 桶装精酿 =====
    # 商品分类为招牌原浆精酿或招牌原浆精酿(L)，排除福袋
    barrel_filter = base_df['商品分类'].astype(str).isin(['招牌原浆精酿', '招牌原浆精酿(L)'])
    barrel_filter = barrel_filter & ~base_df['所属套餐'].astype(str).str.contains('福袋', na=False)
    barrel_df = base_df[barrel_filter]
    result['桶装精酿'] = int(barrel_df['出品数量'].sum())
    
    # ===== 方法2: 瓦猫猫听装精酿 =====
    # 商品分类为招牌精酿瓦猫猫的酒，排除1L规格和二销套餐
    wamao_filter = base_df['商品分类'].astype(str) == '招牌精酿瓦猫猫的酒'
    wamao_filter = wamao_filter & ~base_df['商品名称'].astype(str).str.contains(r'（1L）|\(1L\)', regex=True, na=False)
    wamao_filter = wamao_filter & ~base_df['所属套餐'].astype(str).str.contains('二销', na=False)
    wamao_df = base_df[wamao_filter]
    total_bottles = wamao_df['出品数量'].sum()
    result['瓦猫猫听装精酿'] = round(total_bottles / 12)
    
    # ===== 方法3: 鸡尾酒套餐 =====
    # 商品分类为特调鸡尾酒，商品类型为单品，排除12杯，排除半价(售价=原价一半)
    cocktail_filter = base_df['商品分类'].astype(str) == '特调鸡尾酒'
    cocktail_filter = cocktail_filter & (base_df['商品类型'].astype(str) == '单品')
    cocktail_filter = cocktail_filter & ~base_df['商品名称'].astype(str).str.contains('12杯', na=False)
    
    # 排除半价：商品售价 = 商品原价 / 2
    def is_half_price(row):
        try:
            original_price = float(row.get('商品原价', 0))
            selling_price = float(row.get('商品售价', 0))
            if original_price > 0 and selling_price > 0:
                return abs(selling_price - original_price / 2) < 0.01
            return False
        except:
            return False
    
    cocktail_df = base_df[cocktail_filter]
    cocktail_df = cocktail_df[~cocktail_df.apply(is_half_price, axis=1)]
    result['鸡尾酒套餐'] = int(cocktail_df['出品数量'].sum())
    
    # ===== 方法4: 小吃套餐 =====
    snack_filter = base_df['商品分类'].astype(str) == '小吃套餐'
    snack_df = base_df[snack_filter]
    
    # 59元小吃套餐/小吃A套餐 (正常59元)
    filter_59 = snack_df['商品名称'].astype(str).str.contains('59元小吃套餐|小吃A套餐', na=False)
    filter_59_normal = filter_59 & (snack_df['实付总额'].astype(float) != 49.0)
    result['小吃套餐(59元)'] = int(snack_df[filter_59_normal]['出品数量'].sum())
    
    # 59元小吃套餐/小吃A套餐 (实付49元) - 二销套餐
    filter_59_49 = filter_59 & (snack_df['实付总额'].astype(float) == 49.0)
    result['小吃二销套餐'] = int(snack_df[filter_59_49]['出品数量'].sum())
    
    # 79元小吃套餐/小吃B套餐
    filter_79 = snack_df['商品名称'].astype(str).str.contains('79元小吃套餐|小吃B套餐', na=False)
    result['小吃套餐(79元)'] = int(snack_df[filter_79]['出品数量'].sum())
    
    # 99元小吃套餐/小吃C套餐
    filter_99 = snack_df['商品名称'].astype(str).str.contains('99元小吃套餐|小吃C套餐', na=False)
    result['小吃套餐(99元)'] = int(snack_df[filter_99]['出品数量'].sum())
    
    # ===== 方法5: 1升装精酿双拼套餐 =====
    # 商品分类为招牌精酿瓦猫猫的酒，商品类型为套餐下单品，商品名称包含(1L)
    double_filter = base_df['商品分类'].astype(str) == '招牌精酿瓦猫猫的酒'
    double_filter = double_filter & (base_df['商品类型'].astype(str) == '套餐下单品')
    double_filter = double_filter & base_df['商品名称'].astype(str).str.contains(r'（1L）|\(1L\)', regex=True, na=False)
    double_df = base_df[double_filter]
    result['1升装精酿双拼套餐'] = round(double_df['出品数量'].sum() / 2)
    
    # ===== 方法6: 瓦猫猫二销套餐 =====
    # 商品分类为招牌精酿瓦猫猫的酒，所属套餐包含二销套餐
    erxiao_filter = base_df['商品分类'].astype(str).str.contains('招牌精酿瓦猫猫', na=False)
    erxiao_filter = erxiao_filter & base_df['所属套餐'].astype(str).str.contains('二销套餐', na=False)
    erxiao_df = base_df[erxiao_filter]
    result['瓦猫猫二销套餐'] = round(erxiao_df['出品数量'].sum() / 24)
    
    # ===== 方法7: 小吃二销套餐 =====
    # 59元小吃套餐实付为49元的
    filter_59 = snack_df['商品名称'].astype(str).str.contains('59元小吃套餐|小吃A套餐', na=False)
    filter_49 = filter_59 & (snack_df['实付总额'].astype(float) == 49.0)
    result['小吃二销套餐'] = int(snack_df[filter_49]['出品数量'].sum())
    
    # ===== 方法8: 奔富 =====
    # 商品名称包含奔富，商品类型为单品或套餐下单品
    penfolds_filter = base_df['商品名称'].astype(str).str.contains('奔富', na=False)
    penfolds_filter = penfolds_filter & base_df['商品类型'].astype(str).str.contains('单品|套餐下单品', na=False)
    penfolds_df = base_df[penfolds_filter]
    result['奔富'] = int(penfolds_df['出品数量'].sum())
    
    # ===== 方法9: 点歌 =====
    # 商品名称包含点歌
    diange_filter = base_df['商品名称'].astype(str).str.contains('点歌', na=False)
    diange_df = base_df[diange_filter]
    result['点歌'] = int(diange_df['出品数量'].sum())
    
    # ===== 方法10: 特色斗酒36杯 =====
    # 商品分类为特色斗酒，商品名称包含36杯
    doujiu_filter = base_df['商品分类'].astype(str).str.contains('特色斗酒', na=False)
    doujiu_filter = doujiu_filter & base_df['商品名称'].astype(str).str.contains('36杯', na=False)
    doujiu_df = base_df[doujiu_filter]
    result['特色斗酒36杯'] = int(doujiu_df['出品数量'].sum())
    
    return result


# ============ GUI应用 ============

class CommissionCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("提成计算器 - 门店点餐订单报表分析工具")
        self.root.geometry("1000x700")
        
        self.selected_files = []
        self.results = {}
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI界面"""
        # 标题
        title_label = tk.Label(
            self.root, 
            text="提成计算器", 
            font=("Microsoft YaHei", 24, "bold"),
            fg="#2C3E50"
        )
        title_label.pack(pady=20)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(self.root, text="选择文件", padding=10)
        file_frame.pack(fill="x", padx=20, pady=10)
        
        self.select_btn = ttk.Button(
            file_frame, 
            text="选择CSV文件", 
            command=self.select_files
        )
        self.select_btn.pack(side="left", padx=5)
        
        self.clear_btn = ttk.Button(
            file_frame, 
            text="清空", 
            command=self.clear_files
        )
        self.clear_btn.pack(side="left", padx=5)
        
        self.calc_btn = ttk.Button(
            file_frame, 
            text="计算提成", 
            command=self.calculate,
            state="disabled"
        )
        self.calc_btn.pack(side="left", padx=5)
        
        self.export_btn = ttk.Button(
            file_frame, 
            text="导出Excel", 
            command=self.export_excel,
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=5)
        
        # 文件列表
        self.file_listbox = tk.Listbox(
            self.root, 
            height=8,
            font=("Microsoft YaHei", 10)
        )
        self.file_listbox.pack(fill="x", padx=20, pady=5)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(self.root, text="计算结果", padding=10)
        result_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 创建表格
        columns = ("门店", "桶装精酿", "瓦猫猫听装精酿", "鸡尾酒套餐", 
                   "小吃套餐(59元)", "小吃套餐(79元)", "小吃套餐(99元)", "1升装精酿双拼套餐", "瓦猫猫二销套餐", "小吃二销套餐", "奔富", "点歌", "特色斗酒36杯")
        self.result_tree = ttk.Treeview(
            result_frame, 
            columns=columns, 
            show="headings",
            height=12
        )
        
        # 设置列
        for col in columns:
            self.result_tree.heading(col, text=col)
            self.result_tree.column(col, width=100, anchor="center")
        
        # 滚动条
        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar.set)
        
        self.result_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 状态栏
        self.status_var = tk.StringVar(value="请选择CSV文件")
        self.status_label = tk.Label(
            self.root, 
            textvariable=self.status_var,
            font=("Microsoft YaHei", 9),
            fg="#7F8C8D"
        )
        self.status_label.pack(pady=5)
    
    def select_files(self):
        """选择文件"""
        files = filedialog.askopenfilenames(
            title="选择点餐订单报表CSV文件",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        
        if files:
            self.selected_files = list(files)
            self.file_listbox.delete(0, tk.END)
            
            for f in self.selected_files:
                store_name = extract_store_name(os.path.basename(f))
                self.file_listbox.insert(tk.END, f"{store_name} - {os.path.basename(f)}")
            
            self.status_var.set(f"已选择 {len(files)} 个文件")
            self.calc_btn.config(state="normal")
    
    def clear_files(self):
        """清空文件列表"""
        self.selected_files = []
        self.file_listbox.delete(0, tk.END)
        self.result_tree.delete(*self.result_tree.get_children())
        self.results = {}
        self.status_var.set("请选择CSV文件")
        self.calc_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
    
    def calculate(self):
        """计算提成"""
        if not self.selected_files:
            messagebox.showwarning("警告", "请先选择文件")
            return
        
        self.status_var.set("正在计算...")
        self.result_tree.delete(*self.result_tree.get_children())
        self.results = {}
        
        try:
            for filepath in self.selected_files:
                store_name = extract_store_name(os.path.basename(filepath))
                self.status_var.set(f"正在处理: {store_name}")
                self.root.update()
                
                df = load_csv(filepath)
                result = calculate_commission(df)
                self.results[store_name] = result
                
                # 插入结果行
                self.result_tree.insert("", tk.END, values=(
                    store_name,
                    result.get('桶装精酿', 0),
                    result.get('瓦猫猫听装精酿', 0),
                    result.get('鸡尾酒套餐', 0),
                    result.get('小吃套餐(59元)', 0),
                    result.get('小吃套餐(79元)', 0),
                    result.get('小吃套餐(99元)', 0),
                    result.get('1升装精酿双拼套餐', 0),
                    result.get('瓦猫猫二销套餐', 0),
                    result.get('小吃二销套餐', 0),
                    result.get('奔富', 0),
                    result.get('点歌', 0),
                    result.get('特色斗酒36杯', 0)
                ))
            
            self.status_var.set(f"计算完成，共处理 {len(self.results)} 个门店")
            self.export_btn.config(state="normal")
            messagebox.showinfo("完成", "计算完成！")
            
        except Exception as e:
            self.status_var.set(f"计算失败: {str(e)}")
            messagebox.showerror("错误", f"计算失败: {str(e)}")
    
    def export_excel(self):
        """导出Excel"""
        if not self.results:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="保存Excel报告",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"提成报表_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not filepath:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "提成统计"
            
            # 标题
            ws['A1'] = "门店提成统计报表"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:M1')
            
            # 表头
            headers = ["门店", "桶装精酿", "瓦猫猫听装精酿", "鸡尾酒套餐", 
                      "小吃套餐(59元)", "小吃套餐(79元)", "小吃套餐(99元)", "1升装精酿双拼套餐", "瓦猫猫二销套餐", "小吃二销套餐", "奔富", "点歌", "特色斗酒36杯"]
            for idx, header in enumerate(headers, 1):
                cell = ws.cell(3, idx, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # 数据
            for row_idx, (store_name, data) in enumerate(self.results.items(), 4):
                ws.cell(row_idx, 1, store_name)
                ws.cell(row_idx, 2, data.get('桶装精酿', 0))
                ws.cell(row_idx, 3, data.get('瓦猫猫听装精酿', 0))
                ws.cell(row_idx, 4, data.get('鸡尾酒套餐', 0))
                ws.cell(row_idx, 5, data.get('小吃套餐(59元)', 0))
                ws.cell(row_idx, 6, data.get('小吃套餐(79元)', 0))
                ws.cell(row_idx, 7, data.get('小吃套餐(99元)', 0))
                ws.cell(row_idx, 8, data.get('1升装精酿双拼套餐', 0))
                ws.cell(row_idx, 9, data.get('瓦猫猫二销套餐', 0))
                ws.cell(row_idx, 10, data.get('小吃二销套餐', 0))
                ws.cell(row_idx, 11, data.get('奔富', 0))
                ws.cell(row_idx, 12, data.get('点歌', 0))
                ws.cell(row_idx, 13, data.get('特色斗酒36杯', 0))
            
            # 列宽
            ws.column_dimensions['A'].width = 20
            for col in 'BCDEFGHIJKLM':
                ws.column_dimensions[col].width = 15
            
            wb.save(filepath)
            self.status_var.set(f"已导出: {filepath}")
            messagebox.showinfo("完成", f"Excel报表已保存到:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")


def main():
    root = tk.Tk()
    app = CommissionCalculatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
